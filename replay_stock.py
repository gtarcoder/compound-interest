# -*- coding: utf-8 -*-
import click
import os
import shutil
import pandas as pd
import base64
import openpyxl
import pickle
from datetime import datetime, timedelta
from openpyxl.chart import (
    LineChart,
    Reference,
)
from openpyxl import drawing
from openpyxl.chart.axis import DateAxis

INIT_TOTAL_VALUE = 2000000.0
PER_STOCK_BUY_IN_VALUE = 50000.0
MAX_HOLDING_STOCKS_NUM = 30  # 最多30只自选股
DATE_COLUMN = "A"
TOTAL_VALUE_COLUMN = "B"
TOTAL_PROFIT_RATE_COLUMN = "C"
STOCK_CODE_COLUMN = "D"
STOCK_NAME_COLUMN = "E"
HOLDING_VALUE_COLUMN = "F"
OPENING_PRICE_COLUMN = "G"
CLOSING_PRICE_COLUMN = "H"
BUY_IN_PRICE_COLUMN = "I"
PROFIT_COLUMN = "J"
PROFIT_RATE_COLUMN = "K"
PROFIT_HISTORY_SHEET = "ProfitHistory"
SERIALIZED_INVESTMENT_INFO_SHEET = "SerializedInvestmentInfo"
TOTAL_VALUE_LIST_SIZE = 60
MA_PERIODS = (5, 10, 20, 60)
EARLIEST_DATE = "1971/01/01"


class StockInfo(object):
    __slots__ = ["code", "name"]

    def __init__(self, code, name):
        self.code = code
        self.name = name

    def __str__(self):
        return "code: %s, name: %s" % (self.code, self.name)


class DailyStockPriceInfo(object):
    __slots__ = ["date", "stock_code", "opening_price", "closing_price"]

    def __init__(self, date, stock_code, opening_price, closing_price):
        self.date = date
        self.stock_code = stock_code
        self.opening_price = opening_price
        self.closing_price = closing_price


class HoldingStockInfo(object):
    __slots__ = ["stock", "buy_in_value", "buy_in_price", "current_value",
                 "profit_rate", "current_count", "current_profit", "sell_out_profit", "profit_rate"]

    def __init__(self, stock, buy_in_price, buy_in_value=PER_STOCK_BUY_IN_VALUE):
        self.stock = stock
        self.current_value = self.buy_in_value = buy_in_value
        self.current_count = buy_in_value / buy_in_price  # 股票数量
        self.buy_in_price = buy_in_price
        self.current_profit = self.sell_out_profit = 0

    def calculate_profit(self, cur_price):
        self.current_value = self.current_count * cur_price
        self.current_profit = self.current_value - \
            self.buy_in_value + self.sell_out_profit
        self.profit_rate = self.current_profit / self.buy_in_value

    def buy_in(self, buy_in_price, buy_in_value):
        self.buy_in_value += buy_in_value
        self.current_value += buy_in_value
        self.current_count += buy_in_value / buy_in_price
        self.buy_in_price = (self.buy_in_value + buy_in_value) / \
            (self.buy_in_value / self.buy_in_price + buy_in_value / buy_in_price)

    def sell_out(self, sell_out_price, sell_out_ratio=1.0):
        # 此次卖出的收益
        sell_out_profit = self.current_count * sell_out_ratio * sell_out_price
        # 卖出后剩下的数量和价值
        self.current_count = (1.0 - sell_out_ratio) * self.current_count
        self.current_value = self.current_count * sell_out_price
        self.sell_out_profit += sell_out_profit
        return sell_out_profit

    def __str__(self):
        return """stock: %s, buy_in_price: %.2f, buy_in_value: %s, current_value: %s,
            sell_out_profit: %.2f, current_profit: %.2f, profit_rate: %.2f""" % \
            (self.stock, self.buy_in_price, self.buy_in_value, self.current_value,
             self.sell_out_profit, self.current_profit, 100.0*self.profit_rate) + "%"


class StockMarketHelper(object):
    @staticmethod
    def get_stock_price_info(stock_code, date):
        date = datetime.strptime(date, "%Y/%m/%d").strftime("%Y%m%d")
        url_mode = "http://quotes.money.163.com/service/chddata.html?code=%s&start=%s&end=%s"
        for code_prefix in range(0, 2):
            url = url_mode % ("%d%s" % (
                code_prefix, stock_code), date, date)
            df = pd.read_csv(url, encoding="gb2312")
            if len(df) > 0:
                return DailyStockPriceInfo(date, stock_code, float(df["开盘价"][0]), float(df["收盘价"][0]))
        return None

    @staticmethod
    def is_valid_stock_code(stock_code):
        try:
            int(stock_code)
            return True
        except:
            return False


class InvestmentInfo(object):
    __slots__ = ["holding_stocks", "total_value", "init_value",
                 "cash_value", "stock_value", "profit_rate", "profit_history",
                 "total_value_list", "moving_avg_history", "reduced_stocks", "cur_date"]

    def __init__(self, init_value):
        self.stock_value = 0
        self.cash_value = self.total_value = self.init_value = init_value
        self.holding_stocks = dict()
        self.profit_history = []
        self.total_value_list = [init_value] * TOTAL_VALUE_LIST_SIZE
        self.moving_avg_history = dict()
        self.reduced_stocks = set()
        self.profit_rate = 0
        self.cur_date = EARLIEST_DATE

    def __str__(self):
        return "cur_date: %s, total_value: %.2f, cash_value: %.2f, stock_value: %.2f, profit_rate: %.2f" % \
            (self.cur_date, self.total_value, self.cash_value,
             self.stock_value,
             self.profit_rate*100.0) + "%"

    def above_moving_avg(self, date, period):
        return self.moving_avg_history[date][1] > self.moving_avg_history[date][period]

    def below_moving_avg(self, date, period):
        return self.moving_avg_history[date][1] < self.moving_avg_history[date][period]

    def is_holding_stock(self, stock_code):
        return stock_code in self.holding_stocks

    def get_holding_stock(self, stock_code):
        return self.holding_stocks.get(stock_code)

    def get_holding_stocks(self):
        return [st_code for st_code in self.holding_stocks]

    def calculate_profit(self, date):
        self.stock_value = 0
        for code, stock_info in self.holding_stocks.items():
            self.stock_value += stock_info.current_value

        self.total_value = self.cash_value + self.stock_value
        self.profit_rate = (self.total_value -
                            self.init_value) / self.init_value
        self.profit_history.append((date, self.total_value, self.profit_rate))
        self.total_value_list.append(self.total_value)
        self.total_value_list = self.total_value_list[len(
            self.total_value_list) - TOTAL_VALUE_LIST_SIZE:]

        # 计算均线值
        moving_avg = {1: self.total_value}
        for period in MA_PERIODS:
            moving_avg[period] = sum(self.total_value_list[len(
                self.total_value_list) - period:]) / period
        self.moving_avg_history[date] = moving_avg

    def buy_in_stock(self, stock, date, buy_in_price):
        if stock.code in self.holding_stocks:
            raise RuntimeError(
                "Duplicated holding stock, code %s, name: %s" % (stock.code, stock.name))
        if self.cash_value < PER_STOCK_BUY_IN_VALUE:
            raise RuntimeError(
                "Failed to buy in stock since cash is not enough, current cash: %.2f" % self.cash_value)
        holding_info = HoldingStockInfo(stock, buy_in_price)
        self.holding_stocks[stock.code] = holding_info
        self.cash_value -= PER_STOCK_BUY_IN_VALUE
        print("BUY-IN, stock %s, date %s, buy_in_price %.2f, cash_value %.2f" %
              (stock, date, buy_in_price, self.cash_value))

    def sell_out_stock(self, stock_code, sell_out_date, sell_out_ratio=1.0):
        if stock_code not in self.holding_stocks:
            raise RuntimeError(
                "Failed to sell out stock %s, because stock does not exit in holding stocks" % (stock_code))
        price_info = StockMarketHelper.get_stock_price_info(
            stock_code, sell_out_date)
        if price_info is None:
            raise RuntimeError("Failed to sell out stock %s, failed to get stock price info of date %s" % (
                stock_code, sell_out_date))
        holding_info = self.holding_stocks.get(stock_code)
        # 当天的开盘价卖出
        sell_out_profit = holding_info.sell_out(
            price_info.opening_price, sell_out_ratio)
        # 收益落袋
        self.cash_value += sell_out_profit
        # 计算当前股票收益
        holding_info.calculate_profit(price_info.opening_price)
        print("SELL-OUT, stock info: %s,  date: %s, sell out ratio: %.2f, cash_value: %.2f" %
              (holding_info, sell_out_date, sell_out_ratio, self.cash_value))
        # 删除stock
        if holding_info.current_count == 0:
            del self.holding_stocks[stock_code]


def normalize_date(date):
    if isinstance(date, datetime):
        return date.strftime("%Y/%m/%d")
    elif isinstance(date, str):
        return datetime.strptime(date, "%Y/%m/%d").strftime("%Y/%m/%d")


def get_row_range(sheet, date):
    start_row = -1
    end_row = -1
    for idx in range(2, len(sheet[DATE_COLUMN])):
        if sheet[DATE_COLUMN][idx].value is None:
            continue
        try:
            row_date = normalize_date(sheet[DATE_COLUMN][idx].value)
        except Exception as e:
            row_date = None

        if row_date == date:
            start_row = idx
        if start_row >= 0 and idx > start_row and row_date is not None:
            end_row = idx
            break
    if start_row >= 0 and end_row == -1:
        end_row = len(sheet[DATE_COLUMN])
    return start_row, end_row


def save_stock_info_to_excel(sheet, row, holding_stock_info, price_info):
    sheet["%s%d" % (HOLDING_VALUE_COLUMN, row)
          ] = holding_stock_info.current_value
    sheet["%s%d" % (OPENING_PRICE_COLUMN, row)
          ] = price_info.opening_price
    sheet["%s%d" % (CLOSING_PRICE_COLUMN, row)
          ] = price_info.closing_price
    sheet["%s%d" % (BUY_IN_PRICE_COLUMN, row)
          ] = holding_stock_info.buy_in_price
    sheet["%s%d" % (PROFIT_COLUMN, row)] = holding_stock_info.current_profit
    sheet["%s%d" % (PROFIT_RATE_COLUMN, row)] = "%.2f" % (100.0 *
                                                          holding_stock_info.current_profit / holding_stock_info.buy_in_value) + "%"


def save_investment_info_to_excel(sheet, start_row, end_row, investment_info):
    sheet["%s%d" % (TOTAL_VALUE_COLUMN, start_row + 1)
          ] = investment_info.total_value
    sheet["%s%d" % (TOTAL_PROFIT_RATE_COLUMN, start_row + 1)
          ] = "%.2f" % (100.0*investment_info.profit_rate) + "%"
    sheet.merge_cells("%s%d:%s%d" % (TOTAL_VALUE_COLUMN,
                      start_row + 1, TOTAL_VALUE_COLUMN, end_row))
    sheet.merge_cells("%s%d:%s%d" % (TOTAL_PROFIT_RATE_COLUMN,
                      start_row + 1, TOTAL_PROFIT_RATE_COLUMN, end_row))


def draw_profit_history(wb, investment_info):
    if PROFIT_HISTORY_SHEET in wb.sheetnames:
        wb.remove(wb[PROFIT_HISTORY_SHEET])
    wb.create_sheet(PROFIT_HISTORY_SHEET)
    ws = wb[PROFIT_HISTORY_SHEET]
    ws["A1"] = "日期"
    ws["B1"] = "总资金"
    ws["C1"] = "MA5"
    ws["D1"] = "MA10"
    ws["E1"] = "MA20"
    ws["F1"] = "收益率"
    for idx in range(0, len(investment_info.profit_history)):
        date = investment_info.profit_history[idx][0]
        ws["A%d" %
            (idx + 2)] = datetime.strptime(date, "%Y/%m/%d").date()
        ws["B%d" % (idx + 2)] = investment_info.profit_history[idx][1]
        ws["F%d" % (idx + 2)] = investment_info.profit_history[idx][2]
        ws["C%d" % (idx + 2)] = investment_info.moving_avg_history[date][5]
        ws["D%d" % (idx + 2)] = investment_info.moving_avg_history[date][10]
        ws["E%d" % (idx + 2)] = investment_info.moving_avg_history[date][20]

    total_rows = len(ws["A"])
    asset_line = LineChart()
    asset_line.title = "资金曲线"
    asset_line.style = 12
    asset_line.y_axis.title = "资金"
    asset_line.y_axis.crossAx = 500
    asset_line.x_axis = DateAxis(crossAx=100)
    asset_line.x_axis.title = "日期"
    asset_line.x_axis.number_format = "mm/dd"
    asset_line.x_axis.majorTimeUnit = "days"

    data = Reference(ws, min_col=2, min_row=1, max_col=5,
                     max_row=total_rows)
    asset_line.add_data(data, titles_from_data=True)

    dates = Reference(ws, min_col=1, min_row=2,
                      max_row=total_rows)
    asset_line.set_categories(dates)

    # Style the lines
    s1 = asset_line.series[0]
    s1.graphicalProperties.line = drawing.line.LineProperties(
        solidFill=drawing.colors.ColorChoice(prstClr="green"))

    s2 = asset_line.series[1]
    s2.graphicalProperties.line = drawing.line.LineProperties(
        solidFill=drawing.colors.ColorChoice(prstClr="red"))

    s3 = asset_line.series[2]
    s3.graphicalProperties.line = drawing.line.LineProperties(
        solidFill=drawing.colors.ColorChoice(prstClr="black"))
    s4 = asset_line.series[3]
    s4.graphicalProperties.line = drawing.line.LineProperties(
        solidFill=drawing.colors.ColorChoice(prstClr="blue"))

    ws.add_chart(asset_line, "M20")


def process_daily_stock(investment_info, sheet, cur_date, ref_investment_info=None):
    investment_info.cur_date = cur_date  # 记录当前时间
    start_row, end_row = get_row_range(sheet, cur_date)
    if start_row == -1:
        return
    cur_holding_stocks = []
    for row in range(start_row, end_row):
        stock_code = str(sheet[STOCK_CODE_COLUMN][row].value).strip("\"\".")
        cur_holding_stocks.append(stock_code)

    # 先卖出不在当天自选股名单中的股票
    prev_holding_stocks = investment_info.get_holding_stocks()

    for st_code in prev_holding_stocks:
        if st_code not in cur_holding_stocks:
            # 如果股票从自选股中消失，则按照当天的开盘价卖出
            investment_info.sell_out_stock(st_code, cur_date)

    for row in range(start_row, end_row):
        stock_code = str(sheet[STOCK_CODE_COLUMN][row].value).strip("\"\".")
        stock_name = sheet[STOCK_NAME_COLUMN][row].value.strip("\"\".")
        # 当天的自选股信息
        stock = StockInfo(stock_code, stock_name)
        price_info = StockMarketHelper.get_stock_price_info(
            stock_code, cur_date)
        if price_info is None:
            print("Get empty stock info of stock %s, %s, at date %s" %
                  (stock_code, stock_name, cur_date))
            continue

        # 股票出现在自选股名单中，则按照当天的开盘价买入
        if not investment_info.is_holding_stock(stock_code):
            investment_info.buy_in_stock(
                stock, cur_date, price_info.opening_price)

        holding_stock = investment_info.get_holding_stock(stock_code)

        # 实际账户在 模拟账户当前资金值 < 20日均线值时，股票仓位减半
        if ref_investment_info and ref_investment_info.below_moving_avg(cur_date, 20)\
                and stock_code not in investment_info.reduced_stocks:
            investment_info.sell_out_stock(stock_code, cur_date, 0.5)
            investment_info.reduced_stocks.add(stock_code)

        holding_stock.calculate_profit(price_info.closing_price)
        save_stock_info_to_excel(sheet, row + 1, holding_stock, price_info)

    # 清除减仓的股票记录
    if ref_investment_info and ref_investment_info.above_moving_avg(cur_date, 20):
        investment_info.reduced_stocks.clear()

    # 计算截止到当天整体的收益率
    investment_info.calculate_profit(cur_date)
    print("========== cur date: %s, investment info : %s ==========" %
          (cur_date, investment_info))
    save_investment_info_to_excel(
        sheet, start_row, end_row, investment_info)


def process_stock_account(investment_info, excel, year, ref_investment_info=None):
    try:
        wb = openpyxl.load_workbook(excel)
    except Exception as e:
        raise RuntimeError("Failed to open excel file %s" % excel)

    try:
        end_date = datetime.now().date()
        date = datetime.strptime(
            max(investment_info.cur_date, "%s/01/01" % year), "%Y/%m/%d").date()
        while date <= end_date:
            sheet_name = "%d-%02d" % (year, date.month)
            try:
                ws = wb[sheet_name]
            except Exception as e:
                date = date + timedelta(days=1)
                continue
            if len(ws[DATE_COLUMN]) <= 2:  # 该sheet 为空
                date = date + timedelta(days=1)
                continue

            process_daily_stock(investment_info, ws,
                                date.strftime("%Y/%m/%d"), ref_investment_info)
            date = date + timedelta(days=1)

        draw_profit_history(wb, investment_info)
        # save serialized investment info to excel
        if SERIALIZED_INVESTMENT_INFO_SHEET not in wb.sheetnames:
            wb.create_sheet(SERIALIZED_INVESTMENT_INFO_SHEET)
        ws = wb[SERIALIZED_INVESTMENT_INFO_SHEET]
        ws["A1"] = base64.b64encode(pickle.dumps(investment_info))
    finally:
        wb.save(excel)


def init_investment_info(excel):
    try:
        wb = openpyxl.load_workbook(excel)
        ws = wb[SERIALIZED_INVESTMENT_INFO_SHEET]
        serialized_content = ws["A1"].value
        info = pickle.loads(base64.b64decode(serialized_content))
        return info
    except Exception as e:
        print("Failed to open excel file %s, error: %s" % (excel, e))
        return InvestmentInfo(INIT_TOTAL_VALUE)


@ click.command()
@ click.option("--excel", default="2021-stocks.xlsx", help="excel file path")
@ click.option("--year", default=2021, help="year of stock info")
def process_stock_info(excel, year):
    mock_account_info = init_investment_info(excel)
    process_stock_account(mock_account_info, excel, year)

    excel_new = ".".join(excel.split(".")[:-1]) + "-new.xlsx"
    real_account_info = init_investment_info(excel_new)
    if real_account_info.cur_date == EARLIEST_DATE:  # 最开始直接从 excel 拷贝得到 excel-new
        shutil.copyfile(excel, excel_new)

    process_stock_account(real_account_info, excel_new,
                          year, mock_account_info)


if __name__ == "__main__":
    process_stock_info()
